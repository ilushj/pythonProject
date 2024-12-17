import openpyxl

# 加载模板文件
template_wb = openpyxl.load_workbook('template.xlsx')

# 加载数据Excel文件
data_wb = openpyxl.load_workbook('11.xlsx')

# 获取数据Sheet
data_sheet1 = data_wb['1-发票基本信息']
data_sheet2 = data_wb['2-发票明细信息']

# 设置数据开始行
start_row = 4

# 设置每个新Excel文件的行数
rows_per_file = 100

# 计算总行数
total_rows = data_sheet1.max_row - start_row + 1

# 计算新Excel文件的数量
num_files = -(-total_rows // rows_per_file)  # 向上取整

# 创建新Excel文件
for i in range(num_files):
    # 创建新Excel文件
    new_wb = openpyxl.load_workbook('template.xlsx')

    # 获取新Sheet
    new_sheet1 = new_wb['1-发票基本信息']
    new_sheet2 = new_wb['2-发票明细信息']

    # 计算当前文件的行数
    rows_in_file = min(rows_per_file, total_rows - i * rows_per_file)

    # 复制数据到新Excel文件
    for j in range(rows_in_file):
        row = start_row + i * rows_per_file + j
        new_sheet1.append([cell.value for cell in data_sheet1[row]])
        new_sheet2.append([cell.value for cell in data_sheet2[row]])

        # 保存新Excel文件
    new_wb.save(f'11_{i+1}.xlsx')