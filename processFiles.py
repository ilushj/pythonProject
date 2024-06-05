import os
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
import warnings
from openpyxl.utils import range_boundaries
from openpyxl.comments import Comment
import warnings
from datetime import datetime, timedelta

# 禁用特定警告

def process_excel_files(directory, output_directory):
    # 忽略特定警告
    warnings.filterwarnings("ignore", message="Data Validation extension is not supported and will be removed")
    warnings.filterwarnings("ignore", message="Cell '.*' is part of a merged range but has a comment.*")
    # 创建一个新的工作簿
    wb_new = Workbook()
    ws_new = wb_new.active

    # 设置新工作表的标题行
    ws_new.append(["姓名", "身份证", f, "生效日期", "工种", "有无社保", "备注"])
    tomorrow_date = datetime.now() + timedelta(days=1)
    tomorrow_date = tomorrow_date.date()  # 将日期转换为日期对象
    # 遍历目录下的所有xlsx文件
    for filename in os.listdir(directory):
        if filename.endswith(".xlsx") and filename != "汇总.xlsx" and filename != "奇点客服系统模板.xlsm":
            file_path = os.path.join(directory, filename)

            wb = load_workbook(file_path)

            # 遍历工作表
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # 删除合并单元格中的批注
                for merged_cell_range in ws.merged_cells.ranges:
                    min_row, min_col, max_row, max_col = range_boundaries(merged_cell_range.coord)
                    for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
                        for cell in row:
                            # 检查单元格是否在合并单元格范围内
                            if cell.coordinate in merged_cell_range:
                                # 如果单元格有批注，则将其删除
                                if cell.comment:
                                    cell.comment = None

                # 增员表处理
                if "增员" in sheet_name:
                    # 寻找标题行的索引
                    name_index = id_index = company_index = date_index = occupation_index = None

                    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                        row = list(row)  # 将元组转换为列表
                        name_index = next((i for i, v in enumerate(row) if "姓名" in v), None)
                        id_index = next((i for i, v in enumerate(row) if "证件号码" in v), None)

                        # 寻找包含公司信息的列索引
                        if row:
                            for i, v in enumerate(row):
                                if "公司" in str(v):
                                    company_index = i
                                    break

                        # 寻找包含生效日期信息的列索引
                        if row:
                            date_index = next((i for i, v in enumerate(row) if "生效日期" in str(v)), None)
                            if date_index is None:
                                date_index = next((i for i, v in enumerate(row) if "增员日期" in str(v)), None)
                                if date_index is None:
                                    date_index = next((i for i, v in enumerate(row) if "上班日期" in str(v)), None)

                        # 寻找包含工种信息的列索引
                        if row:
                            occupation_index = next((i for i, v in enumerate(row) if "工种" in str(v)), None)

                    # 检查是否找到了公司列的索引，如果没有，将其设置为 -1
                    if company_index is None:
                        company_index = -1



                    # 检查是否所有必要的列索引都已找到
                    #if all((name_index, id_index, company_index, date_index, occupation_index)):
                        # 遍历数据行并写入新工作表
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        # 如果姓名列不为空，则写入新工作表
                        if row[name_index]:
                            if date_index is not None and row[date_index]:
                                date_value = row[date_index]
                            else:
                                date_value = tomorrow_date
                            ws_new.append([
                                row[name_index],
                                row[id_index],
                                None,
                                "批增" if "增员" in sheet_name else "减员",
                                date_value,
                                row[occupation_index] if occupation_index is not None else None,
                                None,
                                row[company_index]
                            ])

    # 保存新的工作簿
    output_filename = os.path.join(output_directory, datetime.now().strftime("%Y%m%d") + "_奇点模板.xlsx")
    wb_new.save(output_filename)

# 调用函数并传入目录路径
folder_path = r"d:\保全模板test\太保"
output_dir = r"d:\保全模板test\太保"
process_excel_files(folder_path, output_dir)
