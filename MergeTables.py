import openpyxl
import os
from datetime import datetime


def process_daily_report():
    # 设置今日日期
    today_date = datetime.today().strftime('%Y-%m-%d')

    # 定义文件路径
    daily_file = f"D:/千服日报/{today_date}/{today_date}.xlsx"

    # 检查文件是否存在
    if not os.path.exists(daily_file):
        print(f"文件不存在: {daily_file}")
        return

    # 打开目标文件
    source_wb = openpyxl.load_workbook(daily_file)

    # 打开当前工作簿并获取工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 清空当前工作簿Sheet1从第4行开始的所有数据
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # 处理“批减”sheet
    if "批减" in source_wb.sheetnames:
        source_ws = source_wb["批减"]
        source_last_row = source_ws.max_row
        if source_last_row >= 2:
            for i, row in enumerate(source_ws.iter_rows(min_row=2, max_row=source_last_row, min_col=3, max_col=5),
                                    start=4):
                ws[f'B{i}'] = row[0].value  # 新岗位名称.1
                ws[f'C{i}'] = row[1].value  # 新雇员证件类型
                ws[f'D{i}'] = row[2].value  # 新雇员证件号码

    # 处理“批增”sheet
    if "批增" in source_wb.sheetnames:
        source_ws = source_wb["批增"]
        source_last_row = source_ws.max_row
        if source_last_row >= 2:
            for i, row in enumerate(source_ws.iter_rows(min_row=2, max_row=source_last_row, min_col=1, max_col=7),
                                    start=4):
                ws[f'E{i}'] = row[0].value  # 新职业类别
                ws[f'F{i}'] = row[1].value  # 新岗位名称
                ws[f'G{i}'] = row[2].value  # 新岗位名称.1
                ws[f'H{i}'] = row[3].value  # 新雇员证件类型
                ws[f'I{i}'] = row[4].value  # 新雇员证件号码
                ws[f'J{i}'] = row[6].value  # 用工单位

    # 填充A列的顺序数字，考虑批增表的最后一行
    for i in range(4, ws.max_row + 1):
        ws[f'A{i}'] = i - 3

    # 创建一个新的工作簿
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = "Sheet1"

    # 复制数据到新的工作簿
    for row in ws.iter_rows(values_only=True):
        new_ws.append(row)

    # 另存为文件
    save_directory = f"D:/千服日报/{today_date}/"
    save_path = os.path.join(save_directory, "批量记名人员信息变更模板（雇主AB）.xlsx")
    if not os.path.exists(save_directory):
        os.makedirs(save_directory)
    new_wb.save(save_path)

    print(f"文件已另存为: {save_path}")


# 调用函数
process_daily_report()
