import os

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

# 读取整个Excel文件
file_path = r'd:/1.xlsx'  # 替换为你的Excel文件路径
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# 获取每个业务员的唯一值
salesperson_col = [cell.value for cell in ws['B'][1:]]  # 假设“业务员”在A列，从第2行开始
unique_salespersons = set(salesperson_col)

# 定义数据验证规则（普通派遣客户,未指定），确保显示下拉箭头
# dv = DataValidation(type="list", formula1='"普通派遣客户,收单客户,代企业投保,其他"', showDropDown=False)

# 遍历每个业务员，并将他们的数据保存到新的Excel文件中
for salesperson in unique_salespersons:
    # 创建一个新的工作簿并复制整个工作表结构
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active

    # 复制标题行
    for col in ws.iter_cols(min_row=1, max_row=1, values_only=False):
        for cell in col:
            new_ws[cell.coordinate].value = cell.value

    # 复制该业务员的数据行，并为第四列添加数据验证
    row_idx = 2  # 从第二行开始插入数据
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == salesperson:  # 根据“业务员”列的值过滤
            for cell in row:
                new_ws[cell.coordinate.replace(str(cell.row), str(row_idx))].value = cell.value
            row_idx += 1

    # 为第四列添加数据验证（即“客户性质”列，假设在D列）
 #   new_ws.add_data_validation(dv)
  #  dv.add(f'D2:D{row_idx-1}')  # 为该业务员的所有行添加验证

    # 保存为新的Excel文件
    output_path = os.path.join('D:/test111/', f'{salesperson}_data.xlsx')  # 文件名为业务员名字
    # word_output_path = os.path.join('D:/凭证转换/output/', f'{file_name}.docx')
    new_wb.save(output_path)
