import openpyxl
import requests
import os

# 指定Excel文件路径
excel_file_path = r'D:\PIC\1.xlsx'

# 指定下载图片的目录
download_dir = r'D:\PIC\picture'

# 创建下载目录如果不存在
if not os.path.exists(download_dir):
    os.makedirs(download_dir)

# 读取Excel文件
wb = openpyxl.load_workbook(excel_file_path,data_only=True)

# 获取用户输入的sheet名称
sheet_name = input("请输入sheet名称：")

# 获取选定的sheet
sheet = wb[sheet_name]

# 从H列和I列的第三行开始下载图片
for row in range(3, sheet.max_row + 1):
    name_cell = sheet.cell(row=row, column=1)  # 姓名在第一列

    # 获取姓名，如果cell在合并单元格中，则获取合并区域的第一个单元格的值
    for merge in sheet.merged_cells:
        if name_cell.coordinate in merge:
            name = sheet.cell(row=merge.min_row, column=merge.min_col).value
            break
    else:
        name = name_cell.value

    department = sheet.cell(row=row, column=2).value  # 部门在第二列
    date = sheet.cell(row=row, column=3).value  # 日期在第三列
    time = sheet.cell(row=row, column=4).value  # 时间在第四列

    h_cell = sheet.cell(row=row, column=8)  # H列对应的列号是8
    i_cell = sheet.cell(row=row, column=9)  # I列对应的列号是9

    h_link = h_cell.hyperlink.target if h_cell.hyperlink else None
    i_link = i_cell.hyperlink.target if i_cell.hyperlink else None

    links = [link for link in [h_link, i_link] if link]

    print(f"处理第{row}行...")
    print(f"姓名：{name}, 部门：{department}, 日期：{date}, 时间：{time}")
    print(f"有效链接：{links}")

    for link in links:
        print(f"下载图片：{link}")
        try:
            response = requests.get(link, stream=True)
            print(f"响应状态码：{response.status_code}")
            if response.status_code == 200:
                if isinstance(time, str):  # 确保时间为字符串
                    time = time.replace(':', '-')  # 用'-'替换时间中的冒号
                # 生成新的文件名
                image_name = f"{name}_{department}_{date}_{time}.jpg"
                image_name = image_name.replace(':', '_')  # 再次替换冒号
                image_name = image_name.replace('/', '_')  # 替换斜杠
                image_name = image_name.replace('\\', '_')  # 替换反斜杠
                image_name = image_name.replace('*', '_')  # 替换星号
                image_name = image_name.replace('?', '_')  # 替换问号
                image_name = image_name.replace('"', '_')  # 替换双引号
                image_name = image_name.replace('<', '_')  # 替换小于号
                image_name = image_name.replace('>', '_')  # 替换大于号
                image_name = image_name.replace('|', '_')  # 替换竖线

                # 检查文件是否已经存在
                base, extension = os.path.splitext(image_name)
                i = 1
                while os.path.exists(os.path.join(download_dir, image_name)):
                    image_name = f"{base}_{i}{extension}"
                    i += 1

                with open(os.path.join(download_dir, image_name), 'wb') as f:
                    for chunk in response.iter_content(chunk_size=1024):
                        f.write(chunk)
                print(f"下载成功：{link}，重命名为：{image_name}")
            else:
                print(f"下载失败：{link}")
        except requests.exceptions.RequestException as e:
            print(f"下载错误：{e}")
input("Press Enter to exit...")
