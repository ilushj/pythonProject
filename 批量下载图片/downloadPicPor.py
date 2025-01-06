import openpyxl
import requests
import os
from datetime import datetime
from urllib.parse import urlparse


# 工具函数
def sanitize_filename(filename):
    forbidden_chars = [':', '/', '\\', '*', '?', '"', '<', '>', '|']
    for char in forbidden_chars:
        filename = filename.replace(char, '_')
    return filename


def get_unique_filename(directory, filename):
    base, extension = os.path.splitext(filename)
    i = 1
    while os.path.exists(os.path.join(directory, filename)):
        filename = f"{base}_{i}{extension}"
        i += 1
    return filename


def is_valid_url(url):
    parsed = urlparse(url)
    return all([parsed.scheme, parsed.netloc])


def get_cell_value(sheet, row, column):
    """处理合并单元格的情况"""
    cell = sheet.cell(row=row, column=column)
    for merge in sheet.merged_cells:
        if cell.coordinate in merge:
            return sheet.cell(row=merge.min_row, column=merge.min_col).value
    return cell.value


# 主程序
excel_file_path = r'D:\PIC\1.xlsx'
base_download_dir = r'D:\PIC\picture'
os.makedirs(base_download_dir, exist_ok=True)

# 加载Excel文件
wb = openpyxl.load_workbook(excel_file_path, data_only=True)
print("可用的工作表：", wb.sheetnames)
sheet_name = input("请输入sheet名称：")
sheet = wb[sheet_name]

# 获取用户输入的姓名
target_name = input("请输入要查找的姓名（按回车处理所有）：").strip()

# 获取标题行的列索引（从第二行开始）
headers = {sheet.cell(row=2, column=col).value: col for col in range(1, sheet.max_column + 1)}

# 检查标题列的名称
print("标题列：", headers)

# 遍历Excel行
for row in range(3, sheet.max_row + 1):
    # 获取姓名（处理合并单元格）
    name = get_cell_value(sheet, row, headers.get('姓名', 1)) or "Unknown"

    # 如果姓名匹配或用户未输入目标姓名
    if target_name and name != target_name:
        continue  # 跳过不匹配的行

    # 处理其他列信息
    department = get_cell_value(sheet, row, headers.get('部门', 2)) or "Unknown"
    date = get_cell_value(sheet, row, headers.get('日期', 3)) or "Unknown"
    time = get_cell_value(sheet, row, headers.get('时间', 4)) or "Unknown"

    # 获取超链接
    h_link = sheet.cell(row=row, column=headers.get('图1', 8)).hyperlink
    i_link = sheet.cell(row=row, column=headers.get('图2', 9)).hyperlink

    # 获取有效链接
    links = [link.target for link in (h_link, i_link) if link and is_valid_url(link.target)]

    # 动态设置下载目录
    download_dir = os.path.join(base_download_dir, name) if target_name else base_download_dir
    os.makedirs(download_dir, exist_ok=True)

    # 下载图片
    for link in links:
        try:
            response = requests.get(link, stream=True)
            if response.status_code == 200:
                # 生成文件名
                if isinstance(time, datetime):
                    time = time.strftime('%H-%M-%S')
                image_name = sanitize_filename(f"{name}_{department}_{date}_{time}.jpg")
                image_name = get_unique_filename(download_dir, image_name)

                # 保存图片
                with open(os.path.join(download_dir, image_name), 'wb') as f:
                    for chunk in response.iter_content(1024):
                        f.write(chunk)
                print(f"下载成功：{link} -> {image_name}")
            else:
                print(f"下载失败：{link}，状态码：{response.status_code}")
        except requests.exceptions.RequestException as e:
            print(f"下载错误：{e}")

# 提示完成
input("任务完成！按回车退出...")
