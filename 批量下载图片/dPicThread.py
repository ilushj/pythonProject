import openpyxl
import requests
import os
from datetime import datetime
from urllib.parse import urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed


# 工具函数
def sanitize_filename(filename):
    """
    用于清理文件名中的非法字符，以确保在文件系统中有效。
    会将冒号、斜杠、反斜杠等字符替换为下划线。
    """
    forbidden_chars = [':', '/', '\\', '*', '?', '"', '<', '>', '|']
    for char in forbidden_chars:
        filename = filename.replace(char, '_')
    return filename


def get_unique_filename(directory, filename):
    """
    检查目标目录中是否存在指定文件名，如果文件名已存在则为文件名添加数字后缀。
    以避免文件重名，确保保存时不覆盖现有文件。
    """
    base, extension = os.path.splitext(filename)
    i = 1
    while os.path.exists(os.path.join(directory, filename)):
        filename = f"{base}_{i}{extension}"
        i += 1
    return filename


def is_valid_url(url):
    """
    检查给定的URL是否有效（即是否包含协议和域名）。
    使用`urllib.parse.urlparse`来解析URL并验证其组成部分。
    """
    parsed = urlparse(url)
    return all([parsed.scheme, parsed.netloc])


def get_cell_value(sheet, row, column):
    """
    获取指定单元格的值，同时处理合并单元格的情况。
    如果单元格属于合并单元格，则返回合并区域中最左上角单元格的值。
    """
    cell = sheet.cell(row=row, column=column)
    for merge in sheet.merged_cells:
        if cell.coordinate in merge:
            return sheet.cell(row=merge.min_row, column=merge.min_col).value
    return cell.value


def download_image(link, download_dir, name, department, date, time):
    """
    下载单张图片并保存到指定目录。
    根据提供的姓名、部门、日期、时间生成文件名。
    如果下载成功，则将图片保存为本地文件，否则输出错误信息。
    """
    try:
        # 请求图片链接并获取响应
        response = requests.get(link, stream=True)
        if response.status_code == 200:  # 如果请求成功
            # 格式化时间为字符串
            if isinstance(time, datetime):
                time = time.strftime('%H-%M-%S')
            # 生成文件名并清理非法字符
            image_name = sanitize_filename(f"{name}_{department}_{date}_{time}.jpg")
            # 如果文件名重复，获取唯一文件名
            image_name = get_unique_filename(download_dir, image_name)

            # 保存图片
            with open(os.path.join(download_dir, image_name), 'wb') as f:
                for chunk in response.iter_content(1024):  # 分块写入文件
                    f.write(chunk)
            print(f"下载成功：{link} -> {image_name}")
        else:
            print(f"下载失败：{link}，状态码：{response.status_code}")
    except requests.exceptions.RequestException as e:
        # 处理请求异常
        print(f"下载错误：{e}")


# 主程序
excel_file_path = r'D:\PIC\1.xlsx'  # Excel文件路径
base_download_dir = r'D:\PIC\picture'  # 图片下载的根目录
os.makedirs(base_download_dir, exist_ok=True)  # 创建根目录（如果不存在）

# 加载Excel文件
wb = openpyxl.load_workbook(excel_file_path, data_only=True)  # 读取Excel文件
print("可用的工作表：", wb.sheetnames)  # 打印所有工作表名称

# 获取用户输入的sheet名称
sheet_name = input("请输入sheet名称（按回车遍历所有工作表）：").strip()

# 获取用户输入的姓名
target_name = input("请输入要查找的姓名（按回车处理所有）：").strip()

# 如果用户输入了sheet名称，则只处理该sheet，否则遍历所有sheet
if sheet_name:
    sheet_names = [sheet_name]
else:
    sheet_names = wb.sheetnames  # 遍历所有工作表

# 创建一个线程池，设置最大工作线程数为10
with ThreadPoolExecutor(max_workers=5) as executor:
    futures = []  # 用于存储线程任务的返回值（future）

    # 遍历每个工作表
    for sheet_name in sheet_names:
        print(f"开始处理工作表：{sheet_name}")
        sheet = wb[sheet_name]  # 获取当前工作表

        try:
            # 获取标题行的列索引（从第二行开始）
            headers = {sheet.cell(row=2, column=col).value: col for col in range(1, sheet.max_column + 1)}
            print(f"{sheet_name} 工作表的标题列：", headers)  # 打印标题列信息

            # 遍历Excel的每一行（从第三行开始）
            for row in range(3, sheet.max_row + 1):
                # 获取“姓名”列的值（处理合并单元格）
                name = get_cell_value(sheet, row, headers.get('姓名', 1)) or "Unknown"

                # 如果用户指定了目标姓名且当前行的姓名不匹配，则跳过该行
                if target_name and name != target_name:
                    continue  # 跳过不匹配的行

                # 获取其他列的值
                department = get_cell_value(sheet, row, headers.get('部门', 2)) or "Unknown"
                date = get_cell_value(sheet, row, headers.get('日期', 3)) or "Unknown"
                time = get_cell_value(sheet, row, headers.get('时间', 4)) or "Unknown"

                # 获取图片的超链接（如果有）
                h_link = sheet.cell(row=row, column=headers.get('图1', 8)).hyperlink
                i_link = sheet.cell(row=row, column=headers.get('图2', 9)).hyperlink

                # 筛选出有效的超链接（只保留格式正确的URL）
                links = [link.target for link in (h_link, i_link) if link and is_valid_url(link.target)]

                # 动态设置下载目录（根据姓名创建子目录）
                download_dir = os.path.join(base_download_dir, name) if target_name else base_download_dir
                os.makedirs(download_dir, exist_ok=True)  # 创建目录（如果不存在）

                # 将每个图片下载任务提交到线程池
                for link in links:
                    futures.append(executor.submit(download_image, link, download_dir, name, department, date, time))

        except Exception as e:
            print(f"处理工作表 {sheet_name} 时发生错误：{str(e)}")

    # 等待所有下载任务完成
    for future in as_completed(futures):
        future.result()  # 获取线程的执行结果，确保任务完成，捕获异常

# 提示任务完成
input("任务完成！按回车退出...")
