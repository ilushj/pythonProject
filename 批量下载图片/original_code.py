import os
import re
import openpyxl
import requests
from tqdm import tqdm  # 导入 tqdm 库


def sanitize_filename(filename):
    """清理文件名中的非法字符"""
    return re.sub(r'[<>:"/\\|?*]', '_', filename)


def get_cell_value(sheet, row, column):
    """处理合并单元格的情况"""
    cell = sheet.cell(row=row, column=column)

    # 检查是否在合并单元格区域内
    for merge in sheet.merged_cells:
        if cell.coordinate in merge:
            # 如果是合并单元格，返回合并区域第一个单元格的值
            return sheet.cell(row=merge.min_row, column=merge.min_col).value

    # 如果不是合并单元格，直接返回当前单元格的值
    return cell.value


def process_excel_file(file_path, output_dir, sheet_name, target_name="",update_progress=None):
    """处理 Excel 文件并下载图片"""
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 加载 Excel 文件
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]

    for row in tqdm(range(3, sheet.max_row + 1), desc="Processing rows", unit="row"):
        name = get_cell_value(sheet, row, 1)  # 获取姓名
        if target_name and target_name != name:  # 如果输入了目标姓名且不匹配，则跳过
            continue

        department = sheet.cell(row=row, column=2).value or "Unknown"
        date = sheet.cell(row=row, column=3).value or "Unknown"
        time = sheet.cell(row=row, column=4).value or "Unknown"
        h_link = sheet.cell(row=row, column=8).hyperlink
        i_link = sheet.cell(row=row, column=9).hyperlink

        if update_progress:
            progress = (row / sheet.max_row) * 100  # 计算进度百分比
            update_progress(progress)  # 更新进度条

        links = [link.target for link in [h_link, i_link] if link]
        for link in links:
            try:
                response = requests.get(link, stream=True)
                if response.status_code == 200:
                    image_name = f"{name}_{department}_{date}_{time}.jpg"
                    image_name = sanitize_filename(image_name)
                    image_path = os.path.join(output_dir, image_name)
                    with open(image_path, "wb") as f:
                        for chunk in tqdm(response.iter_content(chunk_size=1024), desc=f"Downloading {image_name}", unit="KB"):
                            f.write(chunk)
            except requests.exceptions.RequestException as e:
                print(f"下载错误：{e}")
