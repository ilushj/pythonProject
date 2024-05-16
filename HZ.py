import os
from openpyxl import load_workbook
from openpyxl import Workbook


def find_missing_info(file_path):
    missing_info = []

    # 寻找增员、减员工作表
    wb = load_workbook(file_path)
    has_add_sheet = "增员" in wb.sheetnames
    has_reduce_sheet = "减员" in wb.sheetnames

    # 如果文件中缺少增员或减员工作表，则直接标记并返回
    if not has_add_sheet:
        missing_info.append(("增员", ["文件缺少增员工作表"]))
    if not has_reduce_sheet:
        missing_info.append(("减员", ["文件缺少减员工作表"]))

    for sheet_name in ["增员", "减员"]:
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if sheet.max_row > 1:  # 检查是否有数据行
                # 获取列索引
                name_column_index = None
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=1, column=col).value
                    if cell_value and "姓名" in cell_value:
                        name_column_index = col
                        break

                if name_column_index:
                    # 删除所有姓名列为空的行
                    rows_to_delete = []
                    for row in range(2, sheet.max_row + 1):
                        if not sheet.cell(row=row, column=name_column_index).value:
                            rows_to_delete.append(row)

                    for row in sorted(rows_to_delete, reverse=True):
                        sheet.delete_rows(row, 1)

                # 寻找缺失的字段
                required_fields = ["姓名", "证件号码", "公司"]
                date_fields = []
                if sheet_name == "增员":
                    required_fields.append("工种")
                    date_fields.extend(["生效日期", "增员日期", "上班日期"])
                else:
                    date_fields.extend(["生效日期", "减员日期"])

                # 根据过滤后的数据重新检查缺失字段
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=1, column=col).value
                    if cell_value:
                        for field in required_fields[:]:
                            if field in cell_value:
                                required_fields.remove(field)
                        for date_field in date_fields[:]:
                            if date_field in cell_value:
                                date_fields.remove(date_field)

                if required_fields:
                    missing_info.append((sheet_name, required_fields))
                elif not any(date_fields):  # 如果不存在任何日期字段，则记录缺失
                    missing_info.append((sheet_name, ["日期"]))

    wb.save(file_path)
    wb.close()

    return missing_info
def main():
    # 创建汇总Excel
    summary_wb = Workbook()
    summary_ws = summary_wb.active
    summary_ws.append(["文件名", "缺失字段", "工作表"])

    # 遍历目录下的xlsx文件
    folder_path = r"d:\保全模板test\太保"

    # 检查是否存在汇总.xlsx文件，如果存在则删除
    summary_file_path = os.path.join(folder_path, "汇总.xlsx")
    if os.path.exists(summary_file_path):
        os.remove(summary_file_path)

    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            missing_info = find_missing_info(file_path)
            for sheet_name, missing_fields in missing_info:
                if missing_fields:
                    summary_ws.append([filename, ", ".join(missing_fields), sheet_name])

    # 保存汇总Excel
    summary_wb.save(summary_file_path)


if __name__ == "__main__":
    main()
