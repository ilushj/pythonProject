import os

import fitz  # PyMuPDF.py
import pytesseract
from PIL import Image
from openpyxl import Workbook

# 指定文件夹路径和PDF文件名
folder_path = 'your_folder_path'
pdf_filename = 'shj.pdf'

# 打开PDF文件
pdf_path = os.path.join(folder_path, pdf_filename)
doc = fitz.open(pdf_path)

# 读取每一页的文本
full_text = ''
for page in doc:
    full_text += page.get_text()

# 使用OCR提取金额
# 这里使用了Pillow库来处理图像，确保已安装
images = []
for page_num in range(len(doc)):
    page = doc.load_page(page_num)
    pix = page.get_pixmap()
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    images.append(img)

# 将图像转换为文本
text = ''
for img in images:
    text += pytesseract.image_to_string(img)

# 从文本中提取金额（这里的提取方法取决于金额的格式和所在位置）
# 假设金额在文本中以 "$" 开头
amount_index = text.find('$')
amount = ''
if amount_index != -1:
    amount = text[amount_index:]

# 创建Excel文件
wb = Workbook()
ws = wb.active

# 在Excel中写入金额
ws['A1'] = 'Premium Amount'
ws['B1'] = amount

# 保存Excel文件
excel_filename = 'premium_amount.xlsx'
excel_path = os.path.join(folder_path, excel_filename)
wb.save(excel_path)

print("Excel文件已生成:", excel_path)
