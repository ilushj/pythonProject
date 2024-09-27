import os
import openpyxl
import pandas as pd


def update_and_delete_column():
    # 获取当前工作目录（即exe文件所在目录）
    # current_dir = os.path.dirname(os.path.abspath(__file__))
    current_dir = os.getcwd()

    # 构造文件路径
    order_file_path = os.path.join(current_dir, '订单业绩平分人.xlsx')
    refund_file_path = os.path.join(current_dir, '欠款平分人业绩.xlsx')

    # 读取excel文件
    df_order = pd.read_excel(order_file_path, sheet_name='订单业绩平分人-查询', header=2)
    df_refund = pd.read_excel(refund_file_path, sheet_name='欠款平分人业绩-查询', header=2)

    # 更新“新-系统订单”工作表
    df_order.loc[df_order['归属类型-业绩归属'] == '团队', '归属人-业绩归属'] += '（团队）'
    df_order.drop(columns='归属类型-业绩归属', inplace=True)  # 删除H列

    # 更新“新-回收欠款”工作表
    df_refund.loc[df_refund['销售业绩归属类型'] == '团队', '销售-业绩归属'] += '（团队）'
    df_refund.drop(columns='销售业绩归属类型', inplace=True)  # 删除H列

    # 构造文件路径
    file_path = os.path.join(current_dir, '数据清洗.xlsx')

    # 检查文件是否存在，如果不存在则创建一个空的 DataFrame 并保存为 Excel 文件
    if not os.path.exists(file_path):
        # 创建一个空的 DataFrame
        pd.DataFrame().to_excel(file_path, index=False)

    # 使用 ExcelWriter 追加模式打开文件
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        df_order.to_excel(writer, sheet_name='新-系统订单', index=False)
        df_refund.to_excel(writer, sheet_name='新-回收欠款', index=False)


def clean_sales_performance():
    # 获取当前工作目录（即exe文件所在目录）
    current_dir = os.getcwd()
    # 构造文件路径
    file_path = os.path.join(current_dir, '数据清洗.xlsx')

    # 读取源数据
    df_orders = pd.read_excel(file_path, sheet_name='新-系统订单')
    df_recovery = pd.read_excel(file_path, sheet_name='新-回收欠款')

    # 删除旧的工作表（如果存在）
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        if '销售业绩清洗' in writer.book.sheetnames:
            del writer.book['销售业绩清洗']

        # 创建新工作表
        df_results = pd.DataFrame(
            columns=["销售归属", "销售日期", "业绩类型", "回款金额", "订单编号", "门店编码", "门店名称", "新老业绩判定",
                     "客户名称"])

        # 处理系统订单
        for index, row in df_orders.iterrows():
            if pd.notna(row['归属人-业绩归属']):
                order_type = ''
                if '下店' in str(row['顾客归属']) and row['新老业绩判定'].startswith('新'):
                    order_type = '下店新'
                elif '下店' in str(row['顾客归属']) and row['新老业绩判定'].startswith('老'):
                    order_type = '下店老'
                elif row['新老业绩判定'].startswith('新'):
                    order_type = '引流新'
                else:
                    order_type = '引流老'

                new_row = {
                    "销售归属": row['归属人-业绩归属'],
                    "销售日期": row['付款日期-付款明细'],
                    "业绩类型": order_type,
                    "回款金额": row['理论回款业绩'],
                    "订单编号": row['订单编号'],
                    "门店编码": row['门店编码'],
                    "门店名称": row['门店名称'],
                    "新老业绩判定": row['新老业绩判定'],
                    "客户名称": row['顾客姓名-顾客订单']
                }
                df_results = pd.concat([df_results, pd.DataFrame([new_row])], ignore_index=True)

        # 处理回收欠款
        for index, row in df_recovery.iterrows():
            if pd.notna(row['顾客订单编号-历史欠款']):
                recovery_type = ''
                if row['新老业绩判定'].startswith('新'):
                    recovery_type = '回收店新' if '下店' in str(row['新老业绩判定']) else '回收引新'
                else:
                    recovery_type = '回收店老' if '下店' in str(row['新老业绩判定']) else '回收引老'

                new_row = {
                    "销售归属": row['销售-业绩归属'],
                    "销售日期": row['回收欠款日期'],
                    "业绩类型": recovery_type,
                    "回款金额": row['理论回款业绩'],
                    "订单编号": row['顾客订单编号-历史欠款'],
                    "门店编码": row['门店编码'],
                    "门店名称": row['门店名称'],
                    "新老业绩判定": row['新老业绩判定'],
                    "客户名称": row['顾客姓名-历史欠款']
                }
                df_results = pd.concat([df_results, pd.DataFrame([new_row])], ignore_index=True)

        # 计算合计
        df_results['合计'] = df_results[['回款金额']].sum(axis=1)

        # 写入新工作表
        df_results.to_excel(writer, sheet_name='销售业绩清洗', index=False)


def generate_xshk():
    # 获取当前工作目录（即exe文件所在目录）
    current_dir = os.getcwd()
    # 构造文件路径
    file_path = os.path.join(current_dir, '数据清洗.xlsx')
    # 读取数据
    df_source = pd.read_excel(file_path, sheet_name='销售业绩清洗')

    # 删除旧的工作表（如果存在）
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        if '新XSHK404' in writer.book.sheetnames:
            del writer.book['新XSHK404']

        # 创建新工作表
        df_results = pd.DataFrame(columns=["归属人", "成交日期", "下店新", "引流新", "2开", "回收欠款", "合计"])

        # 处理数据
        for index, row in df_source.iterrows():
            if "新" in str(row['业绩类型']):
                new_row = {
                    "归属人": row['销售归属'],
                    "成交日期": row['销售日期'],
                    "下店新": 0,
                    "引流新": 0,
                    "2开": "",
                    "回收欠款": "",
                    "合计": 0
                }
                if row['业绩类型'] in ["下店新", "回收店新"]:
                    new_row["下店新"] = row['回款金额']
                elif row['业绩类型'] in ["引流新", "回收引新"]:
                    new_row["引流新"] = row['回款金额']

                df_results = pd.concat([df_results, pd.DataFrame([new_row])], ignore_index=True)

        df_results[['下店新', '回收欠款', '引流新']] = df_results[['下店新', '回收欠款', '引流新']].apply(pd.to_numeric,
                                                                                                          errors='coerce')
        df_results['合计'] = df_results[['下店新', '回收欠款', '引流新']].sum(axis=1)

        # 排序
        df_results.sort_values(by='归属人', inplace=True)

        # 写入新工作表
        df_results.to_excel(writer, sheet_name='新XSHK404', index=False)


def merge_sheets():
    # 获取当前工作目录（即exe文件所在目录）
    current_dir = os.getcwd()
    # 构造文件路径
    file_path = os.path.join(current_dir, '数据清洗.xlsx')

    # 打开工作簿
    workbook = openpyxl.load_workbook(file_path)

    # 删除名为“咨询业绩清洗”的工作表（如果存在）
    if '咨询业绩清洗' in workbook.sheetnames:
        std = workbook['咨询业绩清洗']
        workbook.remove(std)

    # 设置要合并的两个sheet和新的sheet
    ws1 = workbook['新-系统订单']
    ws2 = workbook['新-回收欠款']

    # 创建新的sheet并命名为“咨询业绩清洗”
    ws_new = workbook.create_sheet(title='咨询业绩清洗')

    # 确定数据开始的行数
    start_row1 = 1
    start_row2 = 2

    # 获取第一个sheet和第二个sheet的最后一行
    last_row1 = ws1.max_row
    last_row2 = ws2.max_row

    # 将第一个sheet的数据复制到新的sheet中
    for row in ws1.iter_rows(min_row=start_row1, max_row=last_row1, values_only=True):
        ws_new.append(row)

    # 将第二个sheet的数据复制到新的sheet中
    for row in ws2.iter_rows(min_row=start_row2, max_row=last_row2, values_only=True):
        ws_new.append(row)

    # 保存工作簿
    workbook.save(file_path)


if __name__ == "__main__":
    update_and_delete_column()  # 第一步执行
    clean_sales_performance()
    generate_xshk()
    merge_sheets()
